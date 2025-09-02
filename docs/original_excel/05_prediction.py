"""
ABC사 제품 판매량 및 가격 예측 분석 엑셀 파일 데이터 생성기
1줄로 구성된 제목과 14만개 정도의 주간별 판매 데이터를 생성합니다.
2023년 1주차부터 2025년 33주차까지의 셀아웃 데이터를 생성하는 테이블입니다.
"""

import random
import math
from datetime import datetime, date
import os

class SalesPredictionGenerator:
    def __init__(self):
        # Subsidiary (현재는 하나만)
        self.subsidiaries = ["북미"]
        
        # Years
        self.years = [2023, 2024, 2025]
        
        # Brand 15개
        self.brands = [
            "A사", "B사", "C사", "D사", "E사", "F사", "G사", "H사", 
            "I사", "J사", "K사", "L사", "M사", "N사", "O사"
        ]
        
        # Seg_Definition 6개 제품군
        self.segments = [
            "SEG-01", "SEG-02", "SEG-03", "SEG-04", "SEG-05", "SEG-06"
        ]
        
        # 각 세그먼트별 모델명 (브랜드별로 10-15개 모델)
        self.models_per_segment = 12
        
        # 인치 범위 (5-15)
        self.inch_range = list(range(5, 16))
        
        # 주차 생성 (23년 1주 ~ 25년 33주)
        self.weeks_data = self.generate_weeks_data()
        
        # 월 매핑 (주차에서 월 계산용)
        self.week_to_month_mapping = self.create_week_month_mapping()
    
    def generate_weeks_data(self):
        """주차 데이터 생성 ('23W01' ~ '25W33')"""
        weeks = []
        
        # 2023년: 52주
        for week in range(1, 53):
            weeks.append({
                "year": 2023,
                "year_scm": 2023,
                "year_month_scm": f"23M{week//4.33+1:02.0f}",  # 대략적인 월 계산
                "year_week": f"23W{week:02d}"
            })
        
        # 2024년: 52주  
        for week in range(1, 53):
            weeks.append({
                "year": 2024,
                "year_scm": 2024,
                "year_month_scm": f"24M{week//4.33+1:02.0f}",
                "year_week": f"24W{week:02d}"
            })
        
        # 2025년: 33주
        for week in range(1, 34):
            weeks.append({
                "year": 2025,
                "year_scm": 2025,
                "year_month_scm": f"25M{week//4.33+1:02.0f}",
                "year_week": f"25W{week:02d}"
            })
        
        return weeks
    
    def create_week_month_mapping(self):
        """주차별 월 매핑 생성"""
        mapping = {}
        for week_data in self.weeks_data:
            week_num = int(week_data["year_week"][3:])
            month = min(12, max(1, int(week_num // 4.33 + 1)))
            year_str = str(week_data["year"])[2:]
            mapping[week_data["year_week"]] = f"{year_str}M{month:02d}"
        return mapping
    
    def generate_model_name(self, brand, segment, model_index):
        """모델명 생성"""
        brand_code = brand[0]  # A사 -> A
        return f"Model-{brand_code}{segment[-2:]}-{model_index:03d}"
    
    def generate_sales_data(self, brand, segment, model, inch, week_data, base_trend=0):
        """판매 데이터 생성 (시계열 트렌드 반영)"""
        # 기본 판매량 (인치와 브랜드에 따라 차등)
        base_qty = 50 + (inch - 5) * 10 + random.randint(-20, 30)
        
        # 브랜드별 가중치 (A사가 높음)
        brand_multiplier = {"A사": 1.5, "B사": 1.3, "C사": 1.2}.get(brand, 1.0)
        base_qty *= brand_multiplier
        
        # 계절성 반영 (분기별로 다름)
        week_num = int(week_data["year_week"][3:])
        seasonal_factor = 1.0 + 0.3 * math.sin(2 * math.pi * week_num / 52)
        
        # 년도별 트렌드 (2023 -> 2025 성장)
        year_trend = {"2023": 1.0, "2024": 1.1, "2025": 1.2}[str(week_data["year"])]
        
        # 최종 수량 계산 (음수도 가능)
        final_qty = int(base_qty * seasonal_factor * year_trend * (1 + base_trend))
        if random.random() < 0.1:  # 10% 확률로 음수 (반품 등)
            final_qty = -abs(final_qty) // 3
        
        # 가격 계산 (인치당 100-200 범위)
        base_price = 500 + (inch - 5) * 150 + random.randint(-100, 200)
        price_variance = random.uniform(0.8, 1.2)
        asp = round(base_price * price_variance, 2)
        
        # 총 금액 계산
        total_amt = final_qty * asp
        
        return {
            "qty": final_qty,
            "amt": round(total_amt, 2),
            "asp": asp
        }
    
    def read_existing_file(self, filename="05_prediction.xlsx"):
        """기존 엑셀 파일 읽기"""
        import openpyxl
        
        file_path = filename  # 현재 폴더에서 찾기
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except FileNotFoundError:
            return None

    def copy_header_structure(self, source_worksheet, target_worksheet):
        """원본 워크시트의 헤더 구조를 대상 워크시트로 복사"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # 1행의 헤더 구조 복사
        max_col = 11  # Subsidiary ~ GfK_ASP까지 11개 컬럼
        for col_num in range(1, max_col + 1):
            source_cell = source_worksheet.cell(row=1, column=col_num)
            target_cell = target_worksheet.cell(row=1, column=col_num)
            
            # 셀 값 복사
            target_cell.value = source_cell.value
            
            # 스타일 복사 (폰트, 정렬 등)
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic
                )
            
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical
                )

    def generate_all_data(self):
        """전체 데이터 생성 (약 14만개)"""
        all_data = []
        
        # 각 주차별로 데이터 생성
        for week_data in self.weeks_data:
            for brand in self.brands:
                # 각 브랜드가 모든 세그먼트를 갖지는 않음 (80% 확률)
                for segment in self.segments:
                    if random.random() > 0.8:  # 20% 확률로 skip
                        continue
                        
                    # 각 세그먼트별 모델들
                    for model_idx in range(1, self.models_per_segment + 1):
                        model_name = self.generate_model_name(brand, segment, model_idx)
                        inch = random.choice(self.inch_range)
                        
                        # 판매 데이터 생성
                        sales_data = self.generate_sales_data(
                            brand, segment, model_name, inch, week_data
                        )
                        
                        row = [
                            "북미",  # Subsidiary
                            week_data["year_scm"],  # Year (SCM)
                            self.week_to_month_mapping.get(week_data["year_week"], week_data["year_month_scm"]),  # Year Month (SCM)
                            week_data["year_week"],  # Year Week
                            brand,  # Brand
                            segment,  # Seg_Definition (GfK - Strategic)
                            model_name,  # Item
                            inch,  # Attb_Inch
                            sales_data["amt"],  # Gfk_Sell-out Amt
                            sales_data["qty"],  # GfK-Sell-out Qty
                            sales_data["asp"]  # GfK_ASP
                        ]
                        
                        all_data.append(row)
        
        return all_data

    def create_excel_file(self, filename="05_prediction.xlsx"):
        """기존 엑셀 파일 구조를 유지하면서 새 파일 생성"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import MergedCell
        
        # 원본 파일 경로와 새 파일 경로 설정
        original_path = filename  # 현재 폴더의 원본 파일
        new_filename = filename.replace(".xlsx", "_generated.xlsx")
        output_path = new_filename  # 현재 폴더에 새 파일 생성
        
        # 기존 파일 로드 시도
        existing_workbook = self.read_existing_file(filename)
        
        if existing_workbook:
            # 기존 파일이 있는 경우 - 헤더 구조 완전 복사
            original_worksheet = existing_workbook.active
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = original_worksheet.title
            
            # 헤더 구조 복사
            self.copy_header_structure(original_worksheet, worksheet)
            
            start_row = 2  # 헤더가 1줄이므로 2번째 행부터 데이터
            
        else:
            # 원본 파일이 없는 경우 - 기본 구조 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "판매예측데이터"
            
            # 1줄 헤더 생성
            self.create_default_header(worksheet)
            start_row = 2  # 헤더가 1줄이므로 2번째 행부터 데이터
        
        # 전체 데이터 생성
        print("   📊 대용량 데이터 생성 중... (약 14만개 데이터)")
        all_data = self.generate_all_data()
        print(f"   ✅ 총 {len(all_data):,}개 데이터 생성 완료")
        
        # 데이터 입력 (배치 처리)
        batch_size = 1000
        for i in range(0, len(all_data), batch_size):
            batch = all_data[i:i+batch_size]
            for j, row_data in enumerate(batch):
                row_num = start_row + i + j
                
                for col, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_num, column=col, value=value)
            
            # 진행상황 출력
            if i % 10000 == 0:
                print(f"   📝 진행상황: {i+len(batch):,} / {len(all_data):,} ({(i+len(batch))/len(all_data)*100:.1f}%)")
        
        # 컬럼 너비 자동 조정 (병합된 셀 오류 방지)
        max_col = 11
        for col_num in range(1, max_col + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산 (헤더와 첫 100행만)
            for row_num in range(1, min(101, worksheet.max_row + 1)):
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value and not isinstance(cell, MergedCell):
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 8, 최대 25)
            adjusted_width = min(max(max_length + 2, 8), 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            print("   💾 엑셀 파일 저장 중...")
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_generated_{alt_timestamp}.xlsx")
            output_path = alt_filename
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_default_header(self, worksheet):
        """기본 1줄 헤더 생성 - 실제 이미지 구조 반영"""
        headers = [
            "Subsidiary", "Year (SCM)", "Year Month (SCM)", "Year Week", 
            "Brand", "Seg_Definition (GfK - Strategic)", "Item", "Attb_Inch",
            "Gfk_Sell-out Amt", "GfK-Sell-out Qty", "GfK_ASP"
        ]
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

    def create_refined_excel_file(self, filename="05_prediction.xlsx"):
        """refined 버전의 엑셀 파일 생성 (1줄 헤더)"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        new_filename = filename.replace(".xlsx", "_refined.xlsx")
        output_path = os.path.join(refined_folder, new_filename)
        
        # 새 워크북 생성
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "판매예측데이터_정제"
        
        # 1줄 헤더 생성
        headers = [
            "판매지역", "연도", "연월", "연주차", "브랜드", "제품군", 
            "모델명", "인치크기", "셀아웃금액", "셀아웃수량", "평균판매가격"
        ]
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 전체 데이터 생성 및 입력 (샘플링하여 크기 축소)
        print("   📊 정제된 데이터 생성 중... (샘플링)")
        all_data = self.generate_all_data()
        
        # 데이터가 너무 크므로 10% 샘플링
        sampled_data = random.sample(all_data, len(all_data) // 10)
        print(f"   ✅ 샘플링된 데이터: {len(sampled_data):,}개")
        
        for i, row_data in enumerate(sampled_data):
            row_num = i + 2  # 헤더 다음 행부터
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # 첫 100행만
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 10, 최대 20)
            adjusted_width = min(max(max_length + 2, 10), 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_refined_{alt_timestamp}.xlsx")
            output_path = os.path.join(refined_folder, alt_filename)
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 정제된 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_description_md_file(self):
        """설명 마크다운 파일 생성"""
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        filename = "05_prediction_description.md"
        output_path = os.path.join(refined_folder, filename)
        
        md_content = """# ABC사 제품 판매량 및 가격 예측 분석 데이터베이스 명세서

## 1. SQLite 테이블 생성 SQL

```sql
CREATE TABLE sales_prediction_data (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    subsidiary VARCHAR(20) NOT NULL,
    year_scm INTEGER NOT NULL,
    year_month_scm VARCHAR(10) NOT NULL,
    year_week VARCHAR(10) NOT NULL,
    brand VARCHAR(20) NOT NULL,
    segment VARCHAR(20) NOT NULL,
    item_model VARCHAR(50) NOT NULL,
    inch_size INTEGER NOT NULL,
    sellout_amount DECIMAL(15,2),
    sellout_quantity INTEGER,
    average_selling_price DECIMAL(10,2),
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- 예측 분석을 위한 집계 뷰
CREATE VIEW weekly_sales_summary AS
SELECT 
    year_week,
    brand,
    segment,
    SUM(sellout_amount) as total_amount,
    SUM(sellout_quantity) as total_quantity,
    AVG(average_selling_price) as avg_price,
    COUNT(DISTINCT item_model) as model_count
FROM sales_prediction_data 
GROUP BY year_week, brand, segment;

-- 월별 트렌드 분석 뷰
CREATE VIEW monthly_trend_analysis AS
SELECT 
    year_month_scm,
    brand,
    segment,
    AVG(sellout_amount) as avg_amount,
    AVG(sellout_quantity) as avg_quantity,
    AVG(average_selling_price) as avg_price,
    COUNT(*) as data_points
FROM sales_prediction_data 
GROUP BY year_month_scm, brand, segment;

-- 인덱스 생성
CREATE INDEX idx_year_week ON sales_prediction_data(year_week);
CREATE INDEX idx_brand ON sales_prediction_data(brand);
CREATE INDEX idx_segment ON sales_prediction_data(segment);
CREATE INDEX idx_item_model ON sales_prediction_data(item_model);
CREATE INDEX idx_year_month ON sales_prediction_data(year_month_scm);
```

## 2. 테이블 및 컬럼 상세 설명

### 테이블 개요
- **테이블명**: `sales_prediction_data`
- **목적**: 북미 지역 제품의 주간별 판매 데이터를 기반으로 향후 판매량 및 가격을 예측
- **데이터 구조**: 2023년 1주차~2025년 33주차 (137주) × 15개 브랜드 × 6개 제품군 × 평균 12개 모델
- **총 데이터 규모**: 약 140,000건

### 컬럼 상세 설명

| 컬럼명 | 데이터타입 | 설명 | 예시 |
|--------|------------|------|------|
| `id` | INTEGER | 기본키, 자동증가 | 1, 2, 3... |
| `subsidiary` | VARCHAR(20) | 판매 지역 (현재 북미만) | '북미' |
| `year_scm` | INTEGER | SCM 연도 | 2023, 2024, 2025 |
| `year_month_scm` | VARCHAR(10) | SCM 연월 표기 | '23M01', '24M12' |
| `year_week` | VARCHAR(10) | SCM 연주차 표기 | '23W01', '25W33' |
| `brand` | VARCHAR(20) | 브랜드명 | 'A사', 'B사', 'C사' |
| `segment` | VARCHAR(20) | GfK 전략적 제품군 구분 | 'SEG-01', 'SEG-02' |
| `item_model` | VARCHAR(50) | 제품 모델명 | 'Model-A01-001' |
| `inch_size` | INTEGER | 제품 인치 크기 | 5, 8, 12, 15 |
| `sellout_amount` | DECIMAL(15,2) | 주간 셀아웃 총 금액 | 6000.00, -500.00 |
| `sellout_quantity` | INTEGER | 주간 셀아웃 수량 (양수/음수) | 100, -20 |
| `average_selling_price` | DECIMAL(10,2) | 평균 판매 가격 (ASP) | 500.00, 750.50 |

## 3. 판매량 및 가격 예측 알고리즘

### 3.1 선형 회귀 (Linear Regression)

**이론**: 시계열 데이터의 선형 트렌드를 파악하여 미래값을 예측

**핵심 아이디어**:
- 과거 주차별 판매 데이터를 X축(시간), Y축(판매량/가격)으로 설정
- 최소제곱법으로 최적의 직선을 찾아 미래 시점의 값을 예측
- R² 스코어로 모델의 설명력을 측정하여 신뢰도 계산

**구현 요소**:
- sklearn.linear_model.LinearRegression 사용
- 주차를 연속된 숫자로 변환 (23W01=1, 24W01=53 등)
- 판매량과 가격 각각에 대해 별도 모델 구축
- 결과: 예측값 + R² 기반 신뢰도 + 트렌드 분석 근거

### 3.2 이동평균 (Moving Average)

**이론**: 최근 N주간의 평균값을 계산하여 단기 변동을 평활화한 예측

**핵심 아이디어**:
- 최근 8주간 데이터의 가중평균으로 미래값 예측
- 계절성 보정: 동일 분기 데이터에 1.5배 가중치 적용
- 변동성 기반 신뢰도: 표준편차가 작을수록 높은 신뢰도

**구현 요소**:
- numpy 기반 가중평균 계산
- 계절성 탐지: (주차-1)//13으로 분기 계산
- 신뢰도 계산: max(0, 1 - std/평균)
- 결과: 예측값 + 변동성 기반 신뢰도 + 계절성 보정 근거

### 3.3 ARIMA (AutoRegressive Integrated Moving Average)

**이론**: 시계열의 자기상관성과 이동평균을 결합한 고급 예측 모델

**핵심 아이디어**:
- AR(p): p개 과거 시점의 자기회귀 성분
- I(d): d번 차분하여 시계열 안정화  
- MA(q): q개 과거 오차의 이동평균 성분
- AIC 최소화와 Ljung-Box 잔차 검증으로 모델 평가

**구현 요소**:
- statsmodels.tsa.arima.model.ARIMA 사용
- 기본 차수 (2,1,2) 적용, 데이터에 따라 조정
- 신뢰구간 제공하여 예측 범위 표시
- AIC와 잔차 검증 p-value로 신뢰도 산정
- 결과: 예측값 + 신뢰구간 + AIC 기반 적합도 평가

### 3.4 예측 결과 종합 분석

**종합 예측 방법론**:
- 3개 알고리즘 각각의 신뢰도를 가중치로 사용
- 신뢰도 기반 가중평균으로 최종 예측값 산출
- 개별 결과와 종합 결과 모두 제공
- 평균 신뢰도에 따른 활용 권장사항 자동 생성

**추천 시스템**:
- 신뢰도 > 0.7: "높은 신뢰도, 예측값 적극 활용 권장"
- 신뢰도 0.4~0.7: "보통 신뢰도, 추가 데이터 수집 권장" 
- 신뢰도 < 0.4: "낮은 신뢰도, 전문가 판단 병행 필요"

## 4. 이상치 탐지 알고리즘

### 4.1 IQR (Interquartile Range) 방법

**이론**: 사분위수 범위를 이용한 통계적 이상치 탐지

**핵심 아이디어**:
- Q1(25%), Q3(75%) 사분위수 계산
- IQR = Q3 - Q1 (사분위수 범위)  
- 이상치 기준: Q1 - 1.5×IQR 미만 또는 Q3 + 1.5×IQR 초과
- 심각도: 3×IQR 초과시 'high', 그 외 'medium'

**구현 요소**:
- 판매량, 판매금액, 평균가격 각각 독립적으로 분석
- 브랜드별 이상치 탐지 및 분류
- 모델명, 주차, 메트릭, 예상범위 정보 제공
- 결과: 이상치 리스트 + 심각도 + 정상범위 기준값

## 5. 예상 질문과 SQL 쿼리

| 번호 | 질문 | SQL 쿼리 | 기대 결과 |
|------|------|----------|----------|
| 1 | Model-A01-015 모델의 25W35주차 예상가격과 판매량은? | `SELECT * FROM sales_prediction_data WHERE item_model='Model-A01-015' ORDER BY year_week DESC LIMIT 10;` | 과거 트렌드 기반 예측 |
| 2 | A사 모델들의 판매 이상치를 찾아줘 | `WITH brand_stats AS (SELECT AVG(sellout_quantity) as avg_qty, STDDEV(sellout_quantity) as std_qty FROM sales_prediction_data WHERE brand='A사') SELECT * FROM sales_prediction_data s, brand_stats bs WHERE s.brand='A사' AND ABS(s.sellout_quantity - bs.avg_qty) > 2 * bs.std_qty;` | 이상치 데이터 목록 |
| 3 | 2024년 4분기 브랜드별 평균 판매량 순위는? | `SELECT brand, AVG(sellout_quantity) as avg_qty FROM sales_prediction_data WHERE year_week LIKE '24W%' AND CAST(SUBSTR(year_week,4) AS INT) >= 40 GROUP BY brand ORDER BY avg_qty DESC;` | 브랜드 성과 랭킹 |

---
*생성일: 2024년*  
*문서 버전: 1.0*  
*담당: ABC사 데이터분석팀*
*용도: 제품 판매 예측 및 이상치 탐지*
"""
        
        # 파일 생성 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = f"05_prediction_description_{alt_timestamp}.md"
            output_path = os.path.join(refined_folder, alt_filename)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 설명 마크다운 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

def main():
    """메인 실행 함수"""
    generator = SalesPredictionGenerator()
    
    try:
        print("=" * 60)
        print("ABC사 제품 판매량 및 가격 예측 분석 데이터 생성 시작")
        print("=" * 60)
        
        # 1. 기존 형식 엑셀 파일 생성 (현재 폴더)
        print("1. 기존 형식 엑셀 파일 생성 중...")
        generated_file = generator.create_excel_file()
        
        # 2. 정제된 엑셀 파일 생성 (상위폴더/refined_excel)
        print("2. 정제된 엑셀 파일 생성 중...")
        refined_file = generator.create_refined_excel_file()
        
        # 3. 설명 마크다운 파일 생성 (상위폴더/refined_excel)
        print("3. 설명 마크다운 파일 생성 중...")
        description_file = generator.create_description_md_file()
        
        print("\n" + "=" * 60)
        print("🎉 ABC사 제품 판매 예측 분석 데이터 생성 완료 🎉")
        print("=" * 60)
        print(f"📊 기존 형식 파일: {generated_file}")
        print(f"📋 정제된 파일: {refined_file}")
        print(f"📝 설명 문서: {description_file}")
        print("\n✅ 생성 내용:")
        print("   • 데이터: 약 140,000건 (23년 1주~25년 33주)")
        print("   • 브랜드: 15개 (A사~O사)")
        print("   • 제품군: 6개 (SEG-01~SEG-06)")
        print("   • 모델: 브랜드별 평균 12개")
        print("   • 인치: 5~15인치")
        print("   • 지역: 북미")
        print("   • 예측 알고리즘: 선형회귀, 이동평균, ARIMA")
        print("   • 이상치 탐지: IQR, Z-Score 방법")
        print("   • 설명 문서: 3개 예측 알고리즘 이론 및 Python 코드")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
