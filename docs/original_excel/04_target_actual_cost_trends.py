"""
ABC사 KPI 목표 대비 실적 현황 분석 엑셀 파일 데이터 생성기
1줄로 구성된 제목과 다양한 KPI의 월별 목표/실적 데이터를 생성합니다.
각 KPI별로 12개월 × 2(목표/실적) = 24개 데이터를 생성하는 테이블입니다.
"""

import random
from datetime import datetime
import os

class TargetActualTrendsGenerator:
    def __init__(self):
        # KPI 목록
        self.kpis = [
            "총매출금액현황", "고객만족도지수평가", "고객이용자수현황", "광고클릭집계횟수", 
            "구매전환비율통계", "방문자접속횟수", "상품재고보유현황", "운영비용합계액",
            "이용자수현황", "제품불량발생비율", "회원이탈발생비율"
        ]
        
        # 년도
        self.years = [2024, 2025]
        
        # 월 (1~12)
        self.months = list(range(1, 13))
        
        # 목표실적 구분
        self.target_actual = ["목표", "실적"]
        
        # 품종1 (대분류)
        self.category1 = ["전자", "일용품", "곡물", "음료", "화장품", "의류", "가전", "식품"]
        
        # 품종2 (소분류) - 품종1에 따른 매핑
        self.category2_mapping = {
            "전자": ["태블릿", "스마트폰", "노트북", "이어폰", "카메라"],
            "일용품": ["휴지", "세제", "비누", "샴푸", "칫솔"],
            "곡물": ["쌀", "보리", "밀", "옥수수", "콩"],
            "음료": ["주스", "커피", "차", "탄산음료", "물"],
            "화장품": ["크림", "로션", "파운데이션", "립스틱", "마스카라"],
            "의류": ["셔츠", "바지", "원피스", "코트", "신발"],
            "가전": ["냉장고", "세탁기", "전자레인지", "에어컨", "TV"],
            "식품": ["과자", "라면", "빵", "요구르트", "치킨"]
        }
        
        # KPI별 단위와 값 범위
        self.kpi_config = {
            "총매출금액현황": {"unit": "억원", "range": (50.0, 200.0)},
            "고객만족도지수평가": {"unit": "점", "range": (7.0, 10.0)},
            "고객이용자수현황": {"unit": "만명", "range": (10.0, 50.0)},
            "광고클릭집계횟수": {"unit": "만회", "range": (5.0, 25.0)},
            "구매전환비율통계": {"unit": "%", "range": (2.0, 8.0)},
            "방문자접속횟수": {"unit": "만명", "range": (20.0, 100.0)},
            "상품재고보유현황": {"unit": "억원", "range": (30.0, 150.0)},
            "운영비용합계액": {"unit": "억원", "range": (15.0, 80.0)},
            "이용자수현황": {"unit": "만명", "range": (8.0, 40.0)},
            "제품불량발생비율": {"unit": "%", "range": (0.1, 2.0)},
            "회원이탈발생비율": {"unit": "%", "range": (1.0, 5.0)}
        }
    
    def read_existing_file(self, filename="04_target_actual_cost_trends.xlsx"):
        """기존 엑셀 파일 읽기"""
        import openpyxl
        
        file_path = filename  # 현재 폴더에서 찾기
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook
        except FileNotFoundError:
            return None

    def copy_header_structure(self, source_worksheet, target_worksheet):
        """원본 워크시트의 헤더 구조를 대상 워크시트로 복사"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # 1행의 헤더 구조 복사
        max_col = 8  # KPI, 년도, 월, 목표실적, 품종1, 품종2, 값, 단위
        for col_num in range(1, max_col + 1):
            source_cell = source_worksheet.cell(row=1, column=col_num)
            target_cell = target_worksheet.cell(row=1, column=col_num)
            
            # 셀 값 복사
            target_cell.value = source_cell.value
            
            # 스타일 복사 (폰트, 정렬 등)
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic
                )
            
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical
                )

    def generate_value(self, kpi, target_or_actual):
        """KPI와 목표/실적에 따른 값 생성"""
        config = self.kpi_config[kpi]
        min_val, max_val = config["range"]
        
        base_value = random.uniform(min_val, max_val)
        
        # 실적은 목표보다 약간 다르게 (±20% 범위)
        if target_or_actual == "실적":
            variance = random.uniform(-0.2, 0.2)  # ±20%
            base_value *= (1 + variance)
        
        # 소수점 처리
        if config["unit"] in ["점", "%"]:
            return round(base_value, 2)
        else:
            return round(base_value, 2)

    def generate_data_row(self):
        """데이터 행 생성 - 실제 엑셀 구조에 맞춘 데이터"""
        kpi = random.choice(self.kpis)
        year = random.choice(self.years)
        month = random.choice(self.months)
        target_actual = random.choice(self.target_actual)
        category1 = random.choice(self.category1)
        category2 = random.choice(self.category2_mapping[category1])
        
        # 값과 단위 생성
        value = self.generate_value(kpi, target_actual)
        unit = self.kpi_config[kpi]["unit"]
        
        row = [
            kpi,             # KPI
            year,            # 년도
            month,           # 월
            target_actual,   # 목표실적
            category1,       # 품종1
            category2,       # 품종2
            value,           # 값
            unit             # 단위
        ]
        
        return row

    def generate_systematic_data(self):
        """체계적인 데이터 생성 - 각 KPI별로 12개월 × 2(목표/실적) 데이터"""
        all_data = []
        
        # 2024년과 2025년 데이터 생성
        for year in [2024, 2025]:
            for kpi in self.kpis:
                for month in self.months:
                    for target_actual in self.target_actual:
                        # 동일한 KPI/월에 대해 품종1/품종2 조합 생성
                        category1 = random.choice(self.category1)
                        category2 = random.choice(self.category2_mapping[category1])
                        
                        # 목표값 먼저 생성
                        if target_actual == "목표":
                            target_value = self.generate_value(kpi, "목표")
                        else:
                            # 실적은 목표 기준으로 변동
                            target_value = self.generate_value(kpi, "목표")
                            variance = random.uniform(-0.25, 0.25)  # ±25% 변동
                            target_value *= (1 + variance)
                        
                        value = round(target_value, 2)
                        unit = self.kpi_config[kpi]["unit"]
                        
                        row = [
                            kpi,             # KPI
                            year,            # 년도
                            month,           # 월
                            target_actual,   # 목표실적
                            category1,       # 품종1
                            category2,       # 품종2
                            value,           # 값
                            unit             # 단위
                        ]
                        
                        all_data.append(row)
        
        return all_data

    def create_excel_file(self, filename="04_target_actual_cost_trends.xlsx"):
        """기존 엑셀 파일 구조를 유지하면서 새 파일 생성"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import MergedCell
        
        # 원본 파일 경로와 새 파일 경로 설정
        original_path = filename  # 현재 폴더의 원본 파일
        new_filename = filename.replace(".xlsx", "_generated.xlsx")
        output_path = new_filename  # 현재 폴더에 새 파일 생성
        
        # 기존 파일 로드 시도
        existing_workbook = self.read_existing_file(filename)
        
        if existing_workbook:
            # 기존 파일이 있는 경우 - 헤더 구조 완전 복사
            original_worksheet = existing_workbook.active
            
            # 새 워크북 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = original_worksheet.title
            
            # 헤더 구조 복사
            self.copy_header_structure(original_worksheet, worksheet)
            
            start_row = 2  # 헤더가 1줄이므로 2번째 행부터 데이터
            
        else:
            # 원본 파일이 없는 경우 - 기본 구조 생성
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "KPI목표실적현황"
            
            # 1줄 헤더 생성
            self.create_default_header(worksheet)
            start_row = 2  # 헤더가 1줄이므로 2번째 행부터 데이터
        
        # 체계적인 데이터 생성 (각 KPI별로 12개월 × 2 × 2년)
        all_data = self.generate_systematic_data()
        
        # 데이터 입력
        for i, row_data in enumerate(all_data):
            row_num = start_row + i
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정 (병합된 셀 오류 방지)
        max_col = 8
        for col_num in range(1, max_col + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # 성능을 위해 100행까지만 체크
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value and not isinstance(cell, MergedCell):
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 10, 최대 25)
            adjusted_width = min(max(max_length + 2, 10), 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_generated_{alt_timestamp}.xlsx")
            output_path = alt_filename
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_default_header(self, worksheet):
        """기본 1줄 헤더 생성 - 실제 이미지 구조 반영"""
        headers = [
            "KPI", "년도", "월", "목표실적", "품종1", "품종2", "값", "단위"
        ]
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

    def create_refined_excel_file(self, filename="04_target_actual_cost_trends.xlsx"):
        """refined 버전의 엑셀 파일 생성 (1줄 헤더)"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        new_filename = filename.replace(".xlsx", "_refined.xlsx")
        output_path = os.path.join(refined_folder, new_filename)
        
        # 새 워크북 생성
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "KPI목표실적현황_정제"
        
        # 1줄 헤더 생성
        headers = [
            "KPI명", "연도", "월", "구분", "대분류", "소분류", "수치값", "측정단위"
        ]
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # 체계적인 데이터 생성 및 입력
        all_data = self.generate_systematic_data()
        
        for i, row_data in enumerate(all_data):
            row_num = i + 2  # 헤더 다음 행부터
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col, value=value)
        
        # 컬럼 너비 자동 조정
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            # 각 컬럼의 최대 길이 계산
            for row_num in range(1, min(worksheet.max_row + 1, 100)):  # 성능을 위해 100행까지만 체크
                try:
                    cell = worksheet[f"{column_letter}{row_num}"]
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # 컬럼 너비 설정 (최소 12, 최대 20)
            adjusted_width = min(max(max_length + 2, 12), 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            workbook.save(output_path)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = filename.replace(".xlsx", f"_refined_{alt_timestamp}.xlsx")
            output_path = os.path.join(refined_folder, alt_filename)
            workbook.save(output_path)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 정제된 엑셀 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

    def create_description_md_file(self):
        """설명 마크다운 파일 생성"""
        # refined 폴더 경로 설정 (상위폴더/refined_excel)
        refined_folder = os.path.join("..", "refined_excel")
        os.makedirs(refined_folder, exist_ok=True)
        
        filename = "04_target_actual_cost_trends_description.md"
        output_path = os.path.join(refined_folder, filename)
        
        md_content = """# ABC사 KPI 목표 대비 실적 현황 분석 데이터베이스 명세서

## 1. SQLite 테이블 생성 SQL

```sql
CREATE TABLE kpi_target_actual_trends (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    kpi_name VARCHAR(50) NOT NULL,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    target_or_actual VARCHAR(10) NOT NULL,
    category_large VARCHAR(20) NOT NULL,
    category_small VARCHAR(30) NOT NULL,
    value DECIMAL(12,2) NOT NULL,
    unit VARCHAR(10) NOT NULL,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
);

-- KPI 성과 분석을 위한 뷰
CREATE VIEW monthly_kpi_performance AS
SELECT 
    kpi_name,
    year,
    month,
    category_large,
    category_small,
    MAX(CASE WHEN target_or_actual = '목표' THEN value END) as target_value,
    MAX(CASE WHEN target_or_actual = '실적' THEN value END) as actual_value,
    ROUND(
        (MAX(CASE WHEN target_or_actual = '실적' THEN value END) / 
         MAX(CASE WHEN target_or_actual = '목표' THEN value END) * 100), 2
    ) as achievement_rate,
    unit
FROM kpi_target_actual_trends 
GROUP BY kpi_name, year, month, category_large, category_small, unit;

-- 연간 KPI 요약 뷰
CREATE VIEW yearly_kpi_summary AS
SELECT 
    kpi_name,
    year,
    category_large,
    AVG(CASE WHEN target_or_actual = '목표' THEN value END) as avg_target,
    AVG(CASE WHEN target_or_actual = '실적' THEN value END) as avg_actual,
    COUNT(DISTINCT month) as data_months,
    unit
FROM kpi_target_actual_trends 
GROUP BY kpi_name, year, category_large, unit;

-- 인덱스 생성
CREATE INDEX idx_kpi_name ON kpi_target_actual_trends(kpi_name);
CREATE INDEX idx_year_month ON kpi_target_actual_trends(year, month);
CREATE INDEX idx_target_actual ON kpi_target_actual_trends(target_or_actual);
CREATE INDEX idx_category_large ON kpi_target_actual_trends(category_large);
CREATE INDEX idx_category_small ON kpi_target_actual_trends(category_small);
```

## 2. 테이블 및 컬럼 상세 설명

### 테이블 개요
- **테이블명**: `kpi_target_actual_trends`
- **목적**: ABC사의 다양한 KPI에 대한 월별 목표 대비 실적 현황을 관리하고 분석
- **데이터 구조**: 11개 KPI × 12개월 × 2(목표/실적) × 2년 = 약 528건 데이터
- **분석 기능**: 목표 달성률, 월별/연도별 트렌드, 품종별 성과 비교

### 컬럼 상세 설명

| 컬럼명 | 데이터타입 | 설명 | 예시 |
|--------|------------|------|------|
| `id` | INTEGER | 기본키, 자동증가 | 1, 2, 3... |
| `kpi_name` | VARCHAR(50) | KPI 명칭 | '총매출금액현황', '고객만족도지수평가' |
| `year` | INTEGER | 연도 | 2024, 2025 |
| `month` | INTEGER | 월 (1-12) | 1, 6, 12 |
| `target_or_actual` | VARCHAR(10) | 목표/실적 구분 | '목표', '실적' |
| `category_large` | VARCHAR(20) | 대분류 (품종1) | '전자', '일용품', '곡물', '음료' |
| `category_small` | VARCHAR(30) | 소분류 (품종2) | '태블릿', '휴지', '쌀', '주스' |
| `value` | DECIMAL(12,2) | KPI 수치값 | 144.79, 124.81, 8.5 |
| `unit` | VARCHAR(10) | 측정 단위 | '억원', '%', '만명', '점' |

### KPI별 상세 정보

| KPI명 | 측정단위 | 값 범위 | 설명 |
|-------|----------|---------|------|
| 총매출금액현황 | 억원 | 50~200 | 월별 총 매출액 |
| 고객만족도지수평가 | 점 | 7~10 | 고객 만족도 점수 |
| 고객이용자수현황 | 만명 | 10~50 | 월 활성 고객 수 |
| 광고클릭집계횟수 | 만회 | 5~25 | 월 광고 클릭 수 |
| 구매전환비율통계 | % | 2~8 | 방문자 대비 구매 전환율 |
| 방문자접속횟수 | 만명 | 20~100 | 월 웹사이트 방문자 수 |
| 상품재고보유현황 | 억원 | 30~150 | 월말 기준 재고 금액 |
| 운영비용합계액 | 억원 | 15~80 | 월별 총 운영비용 |
| 이용자수현황 | 만명 | 8~40 | 서비스 이용자 수 |
| 제품불량발생비율 | % | 0.1~2.0 | 생산 제품 중 불량률 |
| 회원이탈발생비율 | % | 1~5 | 월 회원 이탈률 |

## 3. 예상 질문과 SQL 쿼리

### 자주 묻는 질문들과 해당 SQL 쿼리

| 번호 | 질문 | SQL 쿼리 | 예상 결과 |
|------|------|----------|----------|
| 1 | 2024년 1월 전자 태블릿의 총매출금액 실적은 얼마야? | `SELECT value FROM kpi_target_actual_trends WHERE kpi_name='총매출금액현황' AND year=2024 AND month=1 AND target_or_actual='실적' AND category_large='전자' AND category_small='태블릿';` | 144.79 억원 |
| 2 | 2024년 음료 주스의 매출 목표와 실적을 비교해줘 | `SELECT month, target_value, actual_value, achievement_rate FROM monthly_kpi_performance WHERE kpi_name='총매출금액현황' AND year=2024 AND category_large='음료' AND category_small='주스';` | 1월: 목표150 실적144.79 등 |
| 3 | 2024년 1월~6월 동안 곡물(쌀)의 매출 추이를 보여줘 | `SELECT month, actual_value FROM monthly_kpi_performance WHERE kpi_name='총매출금액현황' AND year=2024 AND month BETWEEN 1 AND 6 AND category_large='곡물' AND category_small='쌀';` | 1월: 111.18, 2월: 115.28 등 |
| 4 | 총매출금액현황에서 2024년 2월 실적 상위 3개 품종은? | `SELECT category_large, category_small, value FROM kpi_target_actual_trends WHERE kpi_name='총매출금액현황' AND year=2024 AND month=2 AND target_or_actual='실적' ORDER BY value DESC LIMIT 3;` | 전자-태블릿: 144.79 등 |
| 5 | 2024년 1분기 일용품(휴지)의 목표 대비 실적 달성률은? | `SELECT AVG(achievement_rate) as avg_achievement FROM monthly_kpi_performance WHERE kpi_name='총매출금액현황' AND year=2024 AND month BETWEEN 1 AND 3 AND category_large='일용품' AND category_small='휴지';` | 95.2% |
| 6 | 2024년 전체 매출에서 품종1이 '전자'인 항목의 합계는? | `SELECT SUM(value) FROM kpi_target_actual_trends WHERE kpi_name='총매출금액현황' AND year=2024 AND target_or_actual='실적' AND category_large='전자';` | 1,850.5 억원 |
| 7 | 2024년 7월까지 모든 KPI별 목표와 실적 차이를 요약해줘 | `SELECT kpi_name, SUM(CASE WHEN target_or_actual='목표' THEN value END) as total_target, SUM(CASE WHEN target_or_actual='실적' THEN value END) as total_actual FROM kpi_target_actual_trends WHERE year=2024 AND month <= 7 GROUP BY kpi_name;` | 총매출: 목표 1200, 실적 1150 등 |
| 8 | 이용자수현황 KPI에서 2024년 가장 높은 실적을 기록한 월은? | `SELECT month, MAX(value) as max_value FROM kpi_target_actual_trends WHERE kpi_name='이용자수현황' AND year=2024 AND target_or_actual='실적' GROUP BY month ORDER BY max_value DESC LIMIT 1;` | 7월: 38.5만명 |
| 9 | 품종2 기준으로 2024년 매출액이 가장 큰 카테고리는? | `SELECT category_small, SUM(value) as total_sales FROM kpi_target_actual_trends WHERE kpi_name='총매출금액현황' AND year=2024 AND target_or_actual='실적' GROUP BY category_small ORDER BY total_sales DESC LIMIT 1;` | 태블릿: 980.5 억원 |
| 10 | 총매출금액현황 KPI에서 월별 목표 대비 실적이 가장 낮았던 달은? | `SELECT month, MIN(achievement_rate) as min_rate FROM monthly_kpi_performance WHERE kpi_name='총매출금액현황' AND year=2024 GROUP BY month ORDER BY min_rate ASC LIMIT 1;` | 3월: 85.2% |

## 4. 데이터 활용 가이드

### KPI 성과 분석
```sql
-- 월별 KPI 달성률 분석
SELECT 
    kpi_name,
    month,
    AVG(achievement_rate) as avg_achievement_rate,
    COUNT(*) as data_count
FROM monthly_kpi_performance 
WHERE year = 2024
GROUP BY kpi_name, month
ORDER BY kpi_name, month;
```

### 품종별 성과 비교
```sql
-- 대분류별 연간 성과 요약
SELECT 
    category_large,
    kpi_name,
    AVG(CASE WHEN target_or_actual = '목표' THEN value END) as avg_target,
    AVG(CASE WHEN target_or_actual = '실적' THEN value END) as avg_actual,
    ROUND(AVG(CASE WHEN target_or_actual = '실적' THEN value END) / 
          AVG(CASE WHEN target_or_actual = '목표' THEN value END) * 100, 2) as achievement_rate
FROM kpi_target_actual_trends 
WHERE year = 2024 AND kpi_name = '총매출금액현황'
GROUP BY category_large, kpi_name
ORDER BY achievement_rate DESC;
```

### 트렌드 분석
```sql
-- 월별 성장률 분석
SELECT 
    kpi_name,
    month,
    AVG(value) as monthly_avg,
    LAG(AVG(value)) OVER (PARTITION BY kpi_name ORDER BY month) as prev_month_avg,
    ROUND((AVG(value) - LAG(AVG(value)) OVER (PARTITION BY kpi_name ORDER BY month)) / 
          LAG(AVG(value)) OVER (PARTITION BY kpi_name ORDER BY month) * 100, 2) as growth_rate
FROM kpi_target_actual_trends 
WHERE year = 2024 AND target_or_actual = '실적'
GROUP BY kpi_name, month
ORDER BY kpi_name, month;
```

### 이상치 탐지
```sql
-- 목표 달성률이 비정상적인 데이터 탐지
WITH achievement_stats AS (
    SELECT 
        kpi_name,
        AVG(achievement_rate) as avg_rate,
        STDEV(achievement_rate) as std_rate
    FROM monthly_kpi_performance 
    WHERE year = 2024
    GROUP BY kpi_name
)
SELECT 
    mp.kpi_name,
    mp.month,
    mp.category_large,
    mp.category_small,
    mp.achievement_rate,
    CASE 
        WHEN mp.achievement_rate > (as.avg_rate + 2 * as.std_rate) THEN '이상 높음'
        WHEN mp.achievement_rate < (as.avg_rate - 2 * as.std_rate) THEN '이상 낮음'
        ELSE '정상'
    END as status
FROM monthly_kpi_performance mp
JOIN achievement_stats as ON mp.kpi_name = as.kpi_name
WHERE mp.year = 2024
  AND (mp.achievement_rate > (as.avg_rate + 2 * as.std_rate) 
       OR mp.achievement_rate < (as.avg_rate - 2 * as.std_rate))
ORDER BY mp.kpi_name, mp.month;
```

---
*생성일: 2024년*  
*문서 버전: 1.0*  
*담당: ABC사 경영관리팀*
*용도: KPI 목표 대비 실적 분석 및 경영 성과 관리*
"""
        
        # 파일 생성 (권한 오류 방지)
        try:
            # 기존 파일이 있다면 삭제 시도
            if os.path.exists(output_path):
                os.remove(output_path)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
        except PermissionError:
            # 권한 오류 시 다른 파일명으로 재시도
            alt_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 밀리초 포함
            alt_filename = f"04_target_actual_cost_trends_description_{alt_timestamp}.md"
            output_path = os.path.join(refined_folder, alt_filename)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_content)
            print(f"⚠️  파일명 변경됨 (권한 오류): {output_path}")
        
        print(f"✅ 설명 마크다운 파일이 성공적으로 생성되었습니다: {output_path}")
        return output_path

def main():
    """메인 실행 함수"""
    generator = TargetActualTrendsGenerator()
    
    try:
        print("=" * 60)
        print("ABC사 KPI 목표 대비 실적 현황 분석 데이터 생성 시작")
        print("=" * 60)
        
        # 1. 기존 형식 엑셀 파일 생성 (현재 폴더)
        print("1. 기존 형식 엑셀 파일 생성 중...")
        generated_file = generator.create_excel_file()
        
        # 2. 정제된 엑셀 파일 생성 (상위폴더/refined_excel)
        print("2. 정제된 엑셀 파일 생성 중...")
        refined_file = generator.create_refined_excel_file()
        
        # 3. 설명 마크다운 파일 생성 (상위폴더/refined_excel)
        print("3. 설명 마크다운 파일 생성 중...")
        description_file = generator.create_description_md_file()
        
        print("\n" + "=" * 60)
        print("🎉 ABC사 KPI 목표 대비 실적 현황 데이터 생성 완료 🎉")
        print("=" * 60)
        print(f"📊 기존 형식 파일: {generated_file}")
        print(f"📋 정제된 파일: {refined_file}")
        print(f"📝 설명 문서: {description_file}")
        print("\n✅ 생성 내용:")
        print("   • 데이터: 11개 KPI × 12개월 × 2(목표/실적) × 2년 = 528건")
        print("   • KPI: 총매출금액현황, 고객만족도지수평가 등 11개")
        print("   • 연도: 2024년, 2025년")
        print("   • 형식: 1줄 헤더 + 목표/실적 상세 데이터")
        print("   • 품종: 8개 대분류 × 각 5개 소분류")
        print("   • 단위: 억원, %, 만명, 점, 만회 등")
        print("   • 분석 기능: 목표 달성률, 월별 트렌드, 이상치 탐지")
        print("   • 설명 문서: SQLite DB 스키마 + KPI 분석 쿼리")
        print("=" * 60)
        
    except Exception as e:
        print(f"❌ 오류가 발생했습니다: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
